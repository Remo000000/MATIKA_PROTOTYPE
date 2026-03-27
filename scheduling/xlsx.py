from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

from django.http import HttpRequest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from scheduling.schedule_queryset import lessons_queryset_for_request


@dataclass
class WorkbookBytes:
    wb: Workbook

    def to_bytes(self) -> bytes:
        buff = BytesIO()
        self.wb.save(buff)
        return buff.getvalue()


def build_schedule_workbook(*, request: HttpRequest) -> WorkbookBytes:
    qs = lessons_queryset_for_request(request)
    lessons = list(qs)

    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    header = ["Day", "Period", "Group", "Discipline", "Teacher", "Room"]
    ws.append(header)

    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0B5ED7")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for l in lessons:
        ws.append(
            [
                l.timeslot.get_day_of_week_display(),
                l.timeslot.period,
                l.group.name,
                l.discipline.name,
                l.teacher.user.get_short_name(),
                l.room.name,
            ]
        )

    ws.freeze_panes = "A2"
    for col in ("A", "B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 24

    return WorkbookBytes(wb=wb)
